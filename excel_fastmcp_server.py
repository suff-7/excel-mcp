#!/usr/bin/env python3
"""
FastMCP Excel Server - A FastMCP server for Excel file operations
"""

import asyncio
import json
import logging
import re
import os
import datetime
from typing import Any, Dict, List, Optional, Generator
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell

from fastmcp import FastMCP

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("excel-fastmcp-server")

# Get port from environment variable (Render sets this)
PORT = int(os.environ.get("PORT", 8000))

# Create the FastMCP server instance
mcp = FastMCP("excel-server", host="0.0.0.0", port=PORT)

class ExcelFastMCPServer:
    def __init__(self):
        self.mcp = mcp
        self.setup_tools()

    def setup_tools(self):
        """Register all available tools"""
        
        @self.mcp.tool(name="health_check")
        async def health_check() -> str:
            """Health check endpoint for monitoring."""
            return json.dumps({
                "status": "healthy",
                "server": "excel-fastmcp-server",
                "timestamp": datetime.datetime.now().isoformat(),
                "port": PORT
            }, indent=2)
        
        @self.mcp.tool(name="read_excel")
        async def read_excel(file_path: str, sheet_name: str, range: Optional[str] = None) -> str:
            try:
                # Add validation for file extension
                if not file_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                    raise ValueError("Invalid file format. Only Excel files (.xlsx, .xls, .xlsm) are supported")

                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name)
                
                # Add validation for range format
                if range and not re.match(r'^[A-Z]+[0-9]+:[A-Z]+[0-9]+$|^[A-Z]+[0-9]+$', range):
                    raise ValueError("Invalid range format. Use format like 'A1:B10' or 'A1'")
                
                # Improve empty cell handling
                if range:
                    cells = worksheet[range]
                    try:
                        if isinstance(cells, tuple):
                            data = [[cell.value if cell.value is not None else "" for cell in row] for row in cells]
                        elif isinstance(cells, Cell):
                            data = [[cells.value if cells.value is not None else ""]]
                        else:
                            # Defensive processing for other cell types
                            if hasattr(cells, '__iter__'):
                                data = [[cell.value if cell.value is not None else "" for cell in cells]]
                            else:
                                data = [[str(cells)]]  # Fallback
                    except Exception as range_error:
                        # If something goes wrong in range processing, fall back to basic reading
                        import traceback
                        return json.dumps({
                            "error": f"Range processing error: {str(range_error)}",
                            "error_type": type(range_error).__name__,
                            "traceback": traceback.format_exc(),
                            "file_path": file_path,
                            "sheet_name": sheet_name,
                            "range": range
                        })
                    
                    # Extract range boundaries for proper reporting
                    if ':' in range:
                        start_cell, end_cell = range.split(':')
                        start_col, start_row = coordinate_from_string(start_cell)  # Note: coordinate_from_string returns (col, row)
                        end_col, end_row = coordinate_from_string(end_cell)
                        min_row = start_row
                        max_row = end_row
                        min_col = column_index_from_string(start_col)
                        max_col = column_index_from_string(end_col)
                    else:
                        # Single cell
                        single_col, single_row = coordinate_from_string(range)  # Note: coordinate_from_string returns (col, row)
                        min_row = max_row = single_row
                        min_col = max_col = column_index_from_string(single_col)
                else:
                    # Read all rows without skipping - maintain exact row correspondence
                    min_row = worksheet.min_row
                    max_row = worksheet.max_row
                    min_col = worksheet.min_column
                    max_col = worksheet.max_column
                    data = []
                    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, 
                                                min_col=min_col, max_col=max_col):
                        row_data = [cell.value if cell.value is not None else "" for cell in row]
                        data.append(row_data)  # Don't skip empty rows to maintain row indexing
                
                # Safe calculation of dimensions and total cells
                try:
                    rows_count = len(data) if data else 0
                    cols_count = 0
                    if data and len(data) > 0:
                        first_row = data[0]
                        if hasattr(first_row, '__len__') and not isinstance(first_row, str):
                            cols_count = len(first_row)
                    
                    dimensions_str = f"{rows_count} rows x {cols_count} columns"
                    
                    total_cells = 0
                    if data:
                        for row in data:
                            if hasattr(row, '__len__') and not isinstance(row, (str, int, float)):
                                total_cells += len(row)
                            else:
                                total_cells += 1
                except Exception as calc_error:
                    dimensions_str = "Error calculating dimensions"
                    total_cells = 0
                
                return json.dumps({
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "range_read": range if range else f"A{min_row}:{get_column_letter(max_col)}{max_row}",
                    "data": data,
                    "dimensions": dimensions_str,
                    "total_cells": total_cells,
                    "row_offset": min_row - 1  # Add row offset info for easier indexing
                }, indent=2)
            except Exception as e:
                import traceback
                return json.dumps({
                    "error": str(e),
                    "error_type": type(e).__name__,
                    "traceback": traceback.format_exc(),
                    "file_path": file_path,
                    "sheet_name": sheet_name
                })

        @self.mcp.tool(name="write_excel")
        async def write_excel(file_path: str, sheet_name: str, data: List[List[Any]], 
                            start_cell: str = "A1", preserve_formatting: bool = True) -> str:
            try:
                # Validate file extension
                if not file_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                    raise ValueError("Invalid file format. Only Excel files are supported")

                # Validate start_cell format
                if not re.match(r'^[A-Z]+[0-9]+$', start_cell):
                    raise ValueError("Invalid start_cell format. Use format like 'A1'")

                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name, create_sheet=True)
                
                start_row = worksheet[start_cell].row
                start_col = worksheet[start_cell].column
                
                # Backup existing formatting if needed
                if preserve_formatting:
                    existing_formats = {}
                    for row_idx, row_data in enumerate(data):
                        for col_idx, _ in enumerate(row_data):
                            cell = worksheet.cell(row=start_row + row_idx, column=start_col + col_idx)
                            existing_formats[(row_idx, col_idx)] = {
                                'font': cell.font,
                                'fill': cell.fill,
                                'border': cell.border,
                                'alignment': cell.alignment
                            }

                # Write data with type checking
                for row_idx, row_data in enumerate(data):
                    for col_idx, value in enumerate(row_data):
                        cell = worksheet.cell(row=start_row + row_idx, column=start_col + col_idx)
                        
                        # Handle different data types
                        if isinstance(value, (datetime.date, datetime.datetime)):
                            cell.value = value
                            cell.number_format = 'yyyy-mm-dd'
                        elif isinstance(value, (int, float)):
                            cell.value = value
                            cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
                        else:
                            cell.value = str(value) if value is not None else ""

                        # Restore formatting if needed
                        if preserve_formatting and (row_idx, col_idx) in existing_formats:
                            fmt = existing_formats[(row_idx, col_idx)]
                            cell.font = fmt['font']
                            cell.fill = fmt['fill']
                            cell.border = fmt['border']
                            cell.alignment = fmt['alignment']

                # Auto-adjust column widths
                for col_idx in range(len(data[0]) if data else 0):
                    col_letter = get_column_letter(start_col + col_idx)
                    max_length = 0
                    for row_idx in range(len(data)):
                        cell = worksheet.cell(row=start_row + row_idx, column=start_col + col_idx)
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)  # Cap at 50

                workbook.save(file_path)
                
                return json.dumps({
                    "status": "success",
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "rows_written": len(data),
                    "columns_written": len(data[0]) if data else 0,
                    "start_cell": start_cell,
                    "end_cell": f"{get_column_letter(start_col + len(data[0]) - 1)}{start_row + len(data) - 1}",
                    "cells_written": sum(len(row) for row in data)
                }, indent=2)
            except Exception as e:
                return json.dumps({
                    "error": str(e),
                    "error_type": type(e).__name__,
                    "file_path": file_path,
                    "sheet_name": sheet_name
                })

        @self.mcp.tool(name="create_workbook")
        async def create_workbook(file_path: str, sheet_names: List[str] = None) -> str:
            """Create a new, empty Excel workbook."""
            try:
                if sheet_names is None:
                    sheet_names = ["Sheet1"]
                    
                if Path(file_path).exists():
                    raise FileExistsError(f"File already exists at '{file_path}'. Cannot create new workbook.")
                
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)
                for i, sheet_name in enumerate(sheet_names):
                    workbook.create_sheet(sheet_name, index=i)
                
                workbook.save(file_path)
                return json.dumps({
                    "status": "success", "file_path": file_path, "sheets_created": sheet_names
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="list_sheets")
        async def list_sheets(file_path: str) -> str:
            """List the names of all sheets in an Excel workbook."""
            try:
                workbook, _ = self._get_workbook_and_sheet(file_path)
                return json.dumps({
                    "file_path": file_path, "sheets": workbook.sheetnames
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="autofit_columns")
        async def autofit_columns(file_path: str, sheet_name: str, columns: List[str] = None) -> str:
            """Automatically adjust the width of specified columns to fit the content."""
            try:
                if columns is None:
                    columns = []
                    
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name)

                cols_to_fit = columns
                if not cols_to_fit:
                    # If no columns specified, find all columns that have data
                    cols_to_fit = {get_column_letter(c.column) for c in worksheet.columns for r in c if r.value}

                for col_letter in cols_to_fit:
                    max_length = 0
                    for cell in worksheet[col_letter]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[col_letter].width = adjusted_width

                workbook.save(file_path)
                return json.dumps({
                    "status": "success", "file_path": file_path, "sheet_name": sheet_name,
                    "columns_autofit": list(cols_to_fit)
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="format_range")
        async def format_range(file_path: str, sheet_name: str, start_cell: str, end_cell: str = None,
                             bold: bool = False, italic: bool = False, underline: bool = False,
                             font_size: int = None, font_color: str = None, bg_color: str = None,
                             border_style: str = None, border_color: str = None, number_format: str = None,
                             alignment: str = None, wrap_text: bool = False, merge_cells: bool = False) -> str:
            """Apply comprehensive formatting to a range of cells."""
            try:
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name)
                
                range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
                
                # Create font styling
                font_kwargs = {}
                if bold: font_kwargs['bold'] = bold
                if italic: font_kwargs['italic'] = italic
                if underline: font_kwargs['underline'] = 'single' if underline else None
                if font_size: font_kwargs['size'] = font_size
                if font_color: font_kwargs['color'] = font_color.replace('#', '')
                
                font = Font(**font_kwargs) if font_kwargs else None
                
                # Create fill styling
                fill = PatternFill(start_color=bg_color.replace('#', ''), 
                                 end_color=bg_color.replace('#', ''), 
                                 fill_type='solid') if bg_color else None
                
                # Create border styling
                border = None
                if border_style and border_color:
                    side = Side(style=border_style, color=border_color.replace('#', ''))
                    border = Border(left=side, right=side, top=side, bottom=side)
                
                # Create alignment
                alignment_obj = None
                if alignment or wrap_text:
                    align_kwargs = {}
                    if alignment: align_kwargs['horizontal'] = alignment
                    if wrap_text: align_kwargs['wrap_text'] = wrap_text
                    alignment_obj = Alignment(**align_kwargs)
                
                # Apply formatting to range
                for cell in self._iterate_cells_in_range(worksheet, range_str):
                    if font: cell.font = font
                    if fill: cell.fill = fill
                    if border: cell.border = border
                    if alignment_obj: cell.alignment = alignment_obj
                    if number_format: cell.number_format = number_format
                
                # Handle merging
                if merge_cells and end_cell:
                    worksheet.merge_cells(f"{start_cell}:{end_cell}")
                
                workbook.save(file_path)
                
                formatting_applied = {
                    "bold": bold, "italic": italic, "underline": underline,
                    "font_size": font_size, "font_color": font_color, "bg_color": bg_color,
                    "border_style": border_style, "border_color": border_color,
                    "number_format": number_format, "alignment": alignment,
                    "wrap_text": wrap_text, "merge_cells": merge_cells
                }
                
                return json.dumps({
                    "status": "success", "file_path": file_path, "sheet_name": sheet_name,
                    "range": range_str, "formatting_applied": formatting_applied
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="write_data_to_excel")
        async def write_data_to_excel(file_path: str, sheet_name: str, data: List[List[Any]], 
                                    start_cell: str = "A1") -> str:
            """Write data to Excel worksheet with improved type handling."""
            try:
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name, create_sheet=True)
                
                start_row, start_col = coordinate_from_string(start_cell)
                start_col_idx = column_index_from_string(start_col)
                
                # Write data with enhanced type checking
                for row_idx, row_data in enumerate(data):
                    # Ensure row_data is iterable 
                    if not hasattr(row_data, '__iter__') or isinstance(row_data, str):
                        row_data = [row_data]
                    
                    for col_idx, value in enumerate(row_data):
                        cell = worksheet.cell(row=start_row + row_idx, column=start_col_idx + col_idx)
                        
                        # Enhanced data type handling
                        if isinstance(value, (datetime.date, datetime.datetime)):
                            cell.value = value
                            cell.number_format = 'yyyy-mm-dd hh:mm:ss' if isinstance(value, datetime.datetime) else 'yyyy-mm-dd'
                        elif isinstance(value, bool):
                            cell.value = value
                        elif isinstance(value, (int, float)):
                            cell.value = value
                            if isinstance(value, float) and value % 1 != 0:
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'
                        elif value is None:
                            cell.value = ""
                        else:
                            cell.value = str(value)

                workbook.save(file_path)
                
                # Safe calculations for response
                total_rows = len(data) if data else 0
                total_cols = 0
                if data:
                    first_row = data[0] if data else []
                    if hasattr(first_row, '__len__') and not isinstance(first_row, str):
                        total_cols = len(first_row)
                    else:
                        total_cols = 1
                
                end_row = start_row + total_rows - 1 if total_rows > 0 else start_row
                end_col_idx = start_col_idx + total_cols - 1 if total_cols > 0 else start_col_idx
                end_cell = f"{get_column_letter(end_col_idx)}{end_row}"
                
                # Safe cell count calculation
                total_cells = 0
                for row in data:
                    if hasattr(row, '__len__') and not isinstance(row, str):
                        total_cells += len(row)
                    else:
                        total_cells += 1
                
                return json.dumps({
                    "status": "success",
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "rows_written": total_rows,
                    "columns_written": total_cols,
                    "start_cell": start_cell,
                    "end_cell": end_cell,
                    "range_written": f"{start_cell}:{end_cell}",
                    "cells_written": total_cells
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="read_data_from_excel")
        async def read_data_from_excel(file_path: str, sheet_name: str, start_cell: str = "A1", 
                                     end_cell: str = None, preview_only: bool = False) -> str:
            """Read data from Excel worksheet with enhanced metadata."""
            try:
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name)
                
                # Determine range
                if not end_cell:
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column
                    end_cell = f"{get_column_letter(max_col)}{max_row}"
                
                range_str = f"{start_cell}:{end_cell}"
                cells = worksheet[range_str]
                
                # Extract cell data with metadata
                cell_data = []
                if isinstance(cells, Cell):
                    cells = [[cells]]
                elif not isinstance(cells[0], tuple):
                    cells = [cells]
                
                for row_idx, row in enumerate(cells):
                    row_data = []
                    for col_idx, cell in enumerate(row):
                        cell_info = {
                            "address": cell.coordinate,
                            "value": cell.value,
                            "data_type": str(type(cell.value).__name__),
                            "row": cell.row,
                            "column": cell.column,
                            "has_style": cell.has_style
                        }
                        
                        # Add formatting info if cell has style
                        if cell.has_style:
                            cell_info["formatting"] = {
                                "font_bold": cell.font.bold,
                                "font_size": cell.font.size,
                                "font_color": str(cell.font.color.rgb) if cell.font.color else None,
                                "bg_color": str(cell.fill.start_color.rgb) if cell.fill.start_color else None,
                                "number_format": cell.number_format
                            }
                        
                        row_data.append(cell_info)
                        
                        # Limit preview data
                        if preview_only and len(row_data) >= 5:
                            break
                    
                    cell_data.append(row_data)
                    
                    # Limit preview rows
                    if preview_only and len(cell_data) >= 10:
                        break
                
                result = {
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "range_read": range_str,
                    "cells": cell_data,
                    "total_rows": len(cell_data),
                    "total_columns": len(cell_data[0]) if cell_data else 0,
                    "preview_only": preview_only
                }
                
                return json.dumps(result, indent=2, default=str)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="validate_formula_syntax")
        async def validate_formula_syntax(file_path: str, sheet_name: str, cell: str, formula: str) -> str:
            """Validate Excel formula syntax without applying it."""
            try:
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name)
                
                # Basic formula validation
                if not formula.strip():
                    raise ValueError("Formula cannot be empty")
                
                # Remove leading equals sign for validation
                clean_formula = formula.lstrip('=')
                
                # Basic syntax checks
                if not clean_formula:
                    raise ValueError("Formula cannot be just an equals sign")
                
                # Check for balanced parentheses
                if clean_formula.count('(') != clean_formula.count(')'):
                    raise ValueError("Unbalanced parentheses in formula")
                
                # Check for valid cell references pattern
                cell_ref_pattern = r'[A-Z]+[0-9]+'
                if re.search(r'[A-Z]+[0-9]*:', clean_formula):  # Range references
                    range_pattern = r'[A-Z]+[0-9]+:[A-Z]+[0-9]+'
                    if not re.search(range_pattern, clean_formula):
                        raise ValueError("Invalid range reference in formula")
                
                return json.dumps({
                    "status": "valid",
                    "formula": f"={clean_formula}",
                    "cell": cell,
                    "message": "Formula syntax is valid"
                }, indent=2)
            except Exception as e:
                return json.dumps({
                    "status": "invalid",
                    "error": str(e),
                    "formula": formula,
                    "cell": cell
                })

        @self.mcp.tool(name="create_worksheet")
        async def create_worksheet(file_path: str, sheet_name: str) -> str:
            """Create a new worksheet in an existing workbook."""
            try:
                workbook, _ = self._get_workbook_and_sheet(file_path, create_sheet=True)
                
                if sheet_name in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' already exists")
                
                workbook.create_sheet(sheet_name)
                workbook.save(file_path)
                
                return json.dumps({
                    "status": "success",
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "message": f"Worksheet '{sheet_name}' created successfully"
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="delete_worksheet")
        async def delete_worksheet(file_path: str, sheet_name: str) -> str:
            """Delete worksheet from workbook."""
            try:
                workbook, _ = self._get_workbook_and_sheet(file_path, create_sheet=True)
                
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' does not exist")
                
                if len(workbook.sheetnames) == 1:
                    raise ValueError("Cannot delete the only worksheet in the workbook")
                
                worksheet = workbook[sheet_name]
                workbook.remove(worksheet)
                workbook.save(file_path)
                
                return json.dumps({
                    "status": "success",
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "remaining_sheets": workbook.sheetnames,
                    "message": f"Worksheet '{sheet_name}' deleted successfully"
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})


        @self.mcp.tool(name="get_workbook_metadata")
        async def get_workbook_metadata(file_path: str, include_ranges: bool = False) -> str:
            """Get metadata about workbook including sheets and ranges."""
            try:
                workbook, _ = self._get_workbook_and_sheet(file_path)
                
                metadata = {
                    "file_path": file_path,
                    "sheet_count": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames,
                    "active_sheet": workbook.active.title if workbook.active else None
                }
                
                if include_ranges:
                    sheet_info = {}
                    for sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        sheet_info[sheet_name] = {
                            "max_row": ws.max_row,
                            "max_column": ws.max_column,
                            "data_range": f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
                            "merged_cells": [str(merged_range) for merged_range in ws.merged_cells.ranges],
                            "table_count": len(ws.tables)
                        }
                    metadata["sheets_info"] = sheet_info
                
                return json.dumps(metadata, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})


        @self.mcp.tool(name="find_cell_by_value")
        async def find_cell_by_value(file_path: str, sheet_name: str, search_value: str, 
                                   search_range: str = None, exact_match: bool = True) -> str:
            """Find cells containing a specific value and return their addresses."""
            try:
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name)
                
                matches = []
                
                if search_range:
                    cells = worksheet[search_range]
                    if isinstance(cells, Cell):
                        cells = [[cells]]
                    elif not isinstance(cells[0], tuple):
                        cells = [cells]
                else:
                    # Search entire worksheet
                    min_row = worksheet.min_row
                    max_row = worksheet.max_row
                    min_col = worksheet.min_column
                    max_col = worksheet.max_column
                    cells = worksheet.iter_rows(min_row=min_row, max_row=max_row, 
                                              min_col=min_col, max_col=max_col)
                
                for row in cells:
                    if not isinstance(row, tuple):
                        row = [row]
                    for cell in row:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        
                        if exact_match:
                            match_found = cell_value == search_value
                        else:
                            match_found = search_value.lower() in cell_value.lower()
                        
                        if match_found:
                            matches.append({
                                "cell_address": cell.coordinate,
                                "row": cell.row,
                                "column": cell.column,
                                "column_letter": get_column_letter(cell.column),
                                "value": cell.value,
                                "array_row_index": cell.row - worksheet.min_row,  # 0-based index for arrays
                                "array_col_index": cell.column - worksheet.min_column  # 0-based index for arrays
                            })
                
                return json.dumps({
                    "file_path": file_path,
                    "sheet_name": sheet_name,
                    "search_value": search_value,
                    "search_range": search_range,
                    "exact_match": exact_match,
                    "matches": matches,
                    "total_matches": len(matches)
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="add_formula")
        async def add_formula(file_path: str, sheet_name: str, cell: str, formula: str) -> str:
            """Add an Excel formula to a specific cell."""
            try:
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name)
                
                worksheet[cell] = f"={formula.lstrip('=')}"
                
                workbook.save(file_path)
                return json.dumps({
                    "status": "success", "file_path": file_path, "sheet_name": sheet_name,
                    "cell": cell, "formula_added": f"={formula.lstrip('=')}"
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

        @self.mcp.tool(name="update_single_cell")
        async def update_single_cell(file_path: str, sheet_name: str, cell: str, value: str) -> str:
            """Update a single cell with a specific value."""
            try:
                workbook, worksheet = self._get_workbook_and_sheet(file_path, sheet_name, create_sheet=True)
                
                # Set the cell value directly
                worksheet[cell] = value
                
                workbook.save(file_path)
                return json.dumps({
                    "status": "success", 
                    "file_path": file_path, 
                    "sheet_name": sheet_name,
                    "cell": cell, 
                    "value_set": value,
                    "message": f"Cell {cell} updated to '{value}'"
                }, indent=2)
            except Exception as e:
                return json.dumps({"error": str(e)})

    # --- Helper Methods ---

    def _get_workbook_and_sheet(self, file_path: str, sheet_name: Optional[str] = None, 
                               create_sheet: bool = False, data_only: bool = False) -> tuple[openpyxl.Workbook, Optional[Worksheet]]:
        """
        Loads a workbook and a specific sheet, creating them if necessary.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to get/create
            create_sheet: Whether to create sheet if it doesn't exist
            data_only: Whether to load the workbook with data_only=True (formulas as values)
        """
        try:
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            workbook = None
            try:
                # Only try to load if file exists and has content
                if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                    workbook = openpyxl.load_workbook(file_path, data_only=data_only)
                else:
                    raise FileNotFoundError("File does not exist or is empty")
            except (FileNotFoundError, Exception) as e:
                if create_sheet:
                    # Create a new workbook
                    workbook = openpyxl.Workbook()
                    # Remove the default sheet if we're going to create a specific one
                    if sheet_name and sheet_name != workbook.active.title:
                        workbook.remove(workbook.active)
                else:
                    raise ValueError(f"Error loading workbook: {str(e)}")
            
            worksheet = None
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                elif create_sheet:
                    worksheet = workbook.create_sheet(sheet_name)
                else:
                    raise KeyError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            
            return workbook, worksheet
        except Exception as e:
            raise Exception(f"Error in _get_workbook_and_sheet: {str(e)}")

    def _iterate_cells_in_range(self, worksheet: Worksheet, cell_range: str) -> Generator[Cell, None, None]:
        """A helper to yield each cell in a given range string."""
        try:
            cells = worksheet[cell_range]
            if isinstance(cells, Cell):
                yield cells
            elif isinstance(cells, tuple):
                for row in cells:
                    if isinstance(row, tuple):
                        for cell in row:
                            yield cell
                    else:
                        yield row
            else: # A single row range
                for cell in cells:
                    yield cell
        except Exception as e:
            # Fallback for complex ranges
            if ':' in cell_range:
                start_cell, end_cell = cell_range.split(':')
                start_row, start_col = coordinate_from_string(start_cell)
                end_row, end_col = coordinate_from_string(end_cell)
                start_col_idx = column_index_from_string(start_col)
                end_col_idx = column_index_from_string(end_col)
                
                for row in range(start_row, end_row + 1):
                    for col in range(start_col_idx, end_col_idx + 1):
                        yield worksheet.cell(row=row, column=col)
            else:
                yield worksheet[cell_range]

    def run(self):
        """Run the FastMCP server."""
        try:
            logger.info(f"Starting Excel FastMCP Server on host 0.0.0.0 port {PORT}")
            self.mcp.run(transport="streamable-http")
        except Exception as e:
            logger.error(f"Failed to start server: {e}")
            raise

def main():
    """Main entry point"""
    server = ExcelFastMCPServer()
    server.run()

if __name__ == "__main__":
    main() 